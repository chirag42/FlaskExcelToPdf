from flask import Flask,request, render_template, jsonify, make_response, send_file, session
import xlrd
import pandas as pd
from io import StringIO
import os
import openpyxl
from fpdf import FPDF


Upload = "static/upload"
app = Flask(__name__)
app.config['uploadFolder'] = Upload
app.secret_key = "GARIHCLAPGAN"
global fileName

@app.route("/")
def index():
    return "<h1>hello world</h1>"

@app.route("/upload_csv", methods=["GET","POST"])
def upload_csv():
    if request.method == "POST":
        filesize = request.cookies.get("filesize")
        file = request.files["file"]
        session["fileName"] = file.filename

        print(f"Filesize : {filesize}")
        print(file.content_type)
        if(file.content_type != "application/vnd.ms-excel" and file.content_type != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
            response = make_response(jsonify({"message" : f"{file.filename} invalid."}),404)
        else:
            response = make_response(jsonify({"message" : f"{file.filename} uploaded."}),200)
            file.save(os.path.join(app.config['uploadFolder'],file.filename))
            
            #create a pdf file from excel
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial",size=15)
            
            loc = (os.path.join(app.config['uploadFolder'],file.filename))
            wb = openpyxl.load_workbook(loc)
            
            
            for  sheet in wb.worksheets:
                print(sheet)
                print(sheet.max_row)
                for i in range(sheet.max_row - 1):    
                    #pdf.cell(200, 10, txt= sheet.cell(row=1,column=1).value + " " + sheet.cell(row=1,column=2).value , ln =1, align= "C")
                    temp=""
                    #temp+=sheet.cell(row=1,column=1).value + " " + sheet.cell(row=1,column=2).value + "\n"
                    for j in range(sheet.max_column):
                        if(sheet.cell(row=i+2,column=j+1).value != None):
                            temp+=str(sheet.cell(row=1,column=j+1).value) + " : " + str(sheet.cell(row=i+2,column=j+1).value)
                        if(j != sheet.max_column-1 and temp != ""):
                            temp+="\n"
                    if temp != "":              
                        pdf.multi_cell(w=100,h=10,txt=temp,border=1,align="L")
                        pdf.multi_cell(w=100,h=1,txt="",border=0,align="L")

            pdfPath = os.path.join(app.config['uploadFolder'],file.filename[:len(file.filename)-5])+".pdf"

            pdf.output(pdfPath)
            

        return response
    
    return render_template('index.html')

@app.route("/download", methods=["GET","POST"])
def download_file():
    if request.method == "GET":
        print("test")
        fileName = session.get("fileName")
        pdfPath = os.path.join(app.config['uploadFolder'],fileName[:len(fileName)-5])+".pdf"
        #session.pop("fileName")
        return send_file(pdfPath, as_attachment=True)
        
if __name__ == "__main__":
    app.run()

